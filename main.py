from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.exceptions import RefreshError  # Import for handling authentication exceptions
from google.auth.transport.requests import Request  # Import for refreshing tokens
import pickle, os, json
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook

scopes = ['https://www.googleapis.com/auth/calendar']

def get_calendar_service():
    credentials = None
    if os.path.exists("token.pkl"):
        with open("token.pkl", "rb") as token_file:
            credentials = pickle.load(token_file)
    
    # Refresh the token if expired or revoked
    try:
        if not credentials or not credentials.valid:
            if credentials and credentials.expired and credentials.refresh_token:
                credentials.refresh(Request())  # Use Request to refresh the token
            else:
                flow = InstalledAppFlow.from_client_secrets_file("client_secret.json", scopes=scopes)
                credentials = flow.run_local_server(port=0)
                
            # Save the new credentials
            with open("token.pkl", "wb") as token_file:
                pickle.dump(credentials, token_file)
    except RefreshError:
        print("Token expired or revoked. Deleting token file and re-authenticating...")
        if os.path.exists("token.pkl"):
            os.remove("token.pkl")  # Remove the old token file
        # Re-authenticate
        flow = InstalledAppFlow.from_client_secrets_file("client_secret.json", scopes=scopes)
        credentials = flow.run_local_server(port=0)
        
        # Save the new token for future runs
        with open("token.pkl", "wb") as token_file:
            pickle.dump(credentials, token_file)

    return build("calendar", "v3", credentials=credentials)

# Build the calendar service
service = get_calendar_service()

# Get the list of calendars
result = service.calendarList().list().execute()
calendar_id = result['items'][1]['id']  # Replace with correct calendar index if needed

# Load tutoring rates from JSON file
with open("tutoring_rates.json", "r") as rates_file:
    tutor_rates = json.load(rates_file)

# Define the start and end year for the invoice generation
start_year = 2024
current_year = datetime.now().year
current_month = datetime.now().month

# Dictionary to accumulate invoice data for each tutor
invoice_data = {}
monthly_totals = []  # List to hold the total for each month

# Iterate over each month from start_year to the current month
for year in range(start_year, current_year + 1):
    for month in range(1, 13):
        if year == current_year and month > current_month:
            break  # Stop if month is beyond the current month in the current year

        # Define start and end dates for the month
        start_of_month = datetime(year, month, 1).isoformat() + 'Z'
        if month == 12:
            end_of_month = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_of_month = datetime(year, month + 1, 1) - timedelta(days=1)
        end_of_month = end_of_month.isoformat() + 'Z'

        # Get the events for this specific month
        events_result = service.events().list(
            calendarId=calendar_id,
            timeMin=start_of_month,
            timeMax=end_of_month,
            singleEvents=True,
            orderBy='startTime'
        ).execute()

        events = events_result.get('items', [])

        # Filter events with "tutor" in the summary
        tutor_events = [event for event in events if 'tutor' in event.get('summary', '').lower()]

        # Process each event and calculate totals
        monthly_total = 0  # Variable to accumulate total for the current month
        for event in tutor_events:
            title = event.get('summary', 'No Title')
            location = event.get('location', 'Remote')
            description = event.get('description', 'No Description')
            tutor_name = title.split("tutor")[0].strip()

            # Get event start and end times
            start_time = event['start'].get('dateTime')
            end_time = event['end'].get('dateTime')

            if start_time and end_time:
                start_dt = datetime.fromisoformat(start_time.replace("Z", "+00:00"))
                end_dt = datetime.fromisoformat(end_time.replace("Z", "+00:00"))
                duration_hours = (end_dt - start_dt).total_seconds() / 3600

                try:
                    rate = tutor_rates[tutor_name]
                    total_price = rate * duration_hours
                    monthly_total += total_price  # Add to the monthly total

                    if tutor_name not in invoice_data:
                        invoice_data[tutor_name] = {}

                    month_year_key = f"{year}_{month:02d}"

                    if month_year_key not in invoice_data[tutor_name]:
                        invoice_data[tutor_name][month_year_key] = []

                    invoice_data[tutor_name][month_year_key].append({
                        "date": start_dt.strftime('%Y-%m-%d'),
                        "duration_hours": duration_hours,
                        "rate": rate,
                        "total_price": total_price
                    })
                except KeyError:
                    pass  # Skip if no rate is found for the tutor name

        # Add the total for the current month to the monthly_totals list
        if monthly_total > 0:
            month_name = datetime(year, month, 1).strftime('%B %Y')
            monthly_totals.append([month_name, monthly_total])

# Print the monthly totals list
print("Monthly Totals:")
for total in monthly_totals:
    print(total)

# Create the Invoices directory if it doesn't exist
invoices_folder = "Invoices"
os.makedirs(invoices_folder, exist_ok=True)

# Generate a separate Excel file for each tutor and add worksheets for each month
for tutor_name, monthly_sessions in invoice_data.items():
    filename = os.path.join(invoices_folder, f"Invoice_{tutor_name}.xlsx")

    # Load existing workbook or create a new one
    if os.path.exists(filename):
        workbook = load_workbook(filename)
    else:
        workbook = Workbook()
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']  # Remove default sheet if creating new workbook

    # Create a worksheet for each month in the data
    for month_year_key, sessions in monthly_sessions.items():
        # Format sheet name to "Month Year" (e.g., "January 2024")
        year, month = map(int, month_year_key.split('_'))
        sheet_name = f"{datetime(year, month, 1).strftime('%B %Y')}"

        # Avoid creating duplicate sheets
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)

            # Set up header row
            sheet.append(["Date", "Duration (hours)", "Rate ($)", "Total ($)"])

            # Populate rows with session data
            monthly_total = 0
            for session in sessions:
                sheet.append([
                    session["date"],
                    f"{session['duration_hours']:.2f}",
                    f"${session['rate']:.2f}",
                    f"${session['total_price']:.2f}"
                ])
                monthly_total += session['total_price']

            # Add monthly total at the end
            sheet.append([])
            sheet.append(["", "", "Total for Month:", f"${monthly_total:.2f}"])

    # Save the workbook
    workbook.save(filename)
    #print(f"Invoice updated for {tutor_name}: {filename}")

# This script generates or updates an invoice file for each tutor, with worksheets for each month from the start year to the current month, stored in an "Invoices" folder, and also prints the total for each month.
